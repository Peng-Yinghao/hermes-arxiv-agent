#!/usr/bin/env python3
"""
Batch-classify papers by three dimensions using title + abstract only.
Outputs a JSON file that the LLM agent fills in via cron workflow.

Token optimization:
  - Only uses title + abstract (not full PDF)
  - Batches 15 papers per LLM call
  - Structured JSON output with enum values
  - Classification stored in Excel, never re-classified

Usage (by LLM agent):
  1. python3 classify_papers.py --prepare   → writes classify_input.json
  2. LLM reads classify_input.json, fills categories, writes classify_output.json
  3. python3 classify_papers.py --apply     → writes results to Excel
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
INPUT_PATH = BASE_DIR / "classify_input.json"
OUTPUT_PATH = BASE_DIR / "classify_output.json"

CATEGORIES = {
    "scene": ["个人助手", "代码Agent", "Web Agent", "具身Agent", "多智能体模拟", "通用/其他"],
    "form": ["文本", "向量", "结构化", "参数化", "潜变量", "混合/其他"],
    "mem_type": ["上下文/摘要型", "检索型", "反思经验型", "分层管理型", "结构化图谱型", "参数/潜变量型", "其他"],
}


def load_excel_papers():
    wb = load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Papers"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h) if h is not None else "" for h in header_row]
    idx = {name: i for i, name in enumerate(headers)}

    papers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sid = str(row[idx.get("arxiv_id", 0)] or "").strip()
        if not sid:
            continue
        title = str(row[idx.get("title", 1)] or "").strip()
        abstract = str(row[idx.get("abstract", 4)] or "").strip()
        scene = str(row[idx.get("scene", -3)] or "").strip()
        # Only include papers not yet classified
        if scene and scene not in ("", "None"):
            continue
        papers.append({"arxiv_id": sid, "title": title[:200], "abstract": abstract[:800]})
    return papers, headers, idx


def prepare():
    """Write classify_input.json for LLM agent to fill."""
    papers, _, _ = load_excel_papers()
    if not papers:
        print("[OK] All papers already classified.")
        return

    batch_size = 15
    batches = [papers[i:i+batch_size] for i in range(0, len(papers), batch_size)]

    payload = {
        "instructions": (
            "请为以下论文分类，对每篇论文填入三个维度的标签。\n\n"
            "分类维度：\n"
            "  scene (应用场景): 个人助手, 代码Agent, Web Agent, 具身Agent, 多智能体模拟, 通用/其他\n"
            "  form (记忆形式): 文本, 向量, 结构化, 参数化, 潜变量, 混合/其他\n"
            "  mem_type (记忆类型): 上下文/摘要型, 检索型, 反思经验型, 分层管理型, 结构化图谱型, 参数/潜变量型, 其他\n\n"
            "只根据论文标题和摘要判断，不要猜测。不确定时选最接近的或'其他'。\n"
            "只返回 JSON 数组，不要其他文字。"
        ),
        "batches": [
            {
                "batch_id": i,
                "papers": batch,
                "expected_output": [{"arxiv_id": p["arxiv_id"], "scene": "...", "form": "...", "mem_type": "..."} for p in batch]
            }
            for i, batch in enumerate(batches)
        ],
        "output_format": '[{"arxiv_id":"...","scene":"...","form":"...","mem_type":"..."},...]',
    }

    INPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[OK] {len(papers)} papers to classify in {len(batches)} batches → {INPUT_PATH}")


def apply():
    """Read classify_output.json and write to Excel."""
    if not OUTPUT_PATH.exists():
        print(f"[ERROR] {OUTPUT_PATH} not found. Run classify first.")
        sys.exit(1)

    results = json.loads(OUTPUT_PATH.read_text(encoding="utf-8"))
    if isinstance(results, dict):
        results = results.get("results", results.get("papers", []))

    result_map = {r["arxiv_id"]: r for r in results}

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Papers"]
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h) if h is not None else "" for h in header_row]

    # Ensure classification columns exist
    for col_name in ["scene", "form", "mem_type"]:
        if col_name not in headers:
            col = len(headers) + 1
            ws.cell(row=1, column=col, value=col_name)
            headers.append(col_name)
    idx = {name: i + 1 for i, name in enumerate(headers)}

    updated = 0
    for row in range(2, ws.max_row + 1):
        sid = str(ws.cell(row=row, column=idx["arxiv_id"]).value or "").strip()
        if sid in result_map:
            r = result_map[sid]
            ws.cell(row=row, column=idx["scene"], value=r.get("scene", ""))
            ws.cell(row=row, column=idx["form"], value=r.get("form", ""))
            ws.cell(row=row, column=idx["mem_type"], value=r.get("mem_type", ""))
            updated += 1

    wb.save(EXCEL_PATH)
    print(f"[OK] Updated {updated} papers in Excel")


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 classify_papers.py --prepare | --apply")
        sys.exit(1)

    cmd = sys.argv[1]
    if cmd == "--prepare":
        prepare()
    elif cmd == "--apply":
        apply()
    else:
        print(f"Unknown command: {cmd}")


if __name__ == "__main__":
    main()
