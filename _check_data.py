import openpyxl
from collections import Counter

wb = openpyxl.load_workbook("papers_record.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]
print("Columns:", headers)

if "source" in headers:
    idx = headers.index("source")
    sources = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx]:
            sources[row[idx]] += 1
    print("\nSource distribution:")
    for k,v in sources.most_common():
        print(f"  {k}: {v}")

if "summary_cn" in headers:
    idx_sc = headers.index("summary_cn")
    idx_id = headers.index("arxiv_id") if "arxiv_id" in headers else 0
    missing_sc = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[idx_sc]
        if not val or str(val).strip() == "":
            missing_sc.append(row[idx_id])
    print(f"\nMissing summary_cn: {len(missing_sc)}")
    print(f"Sample IDs: {missing_sc[:8]}")

print(f"\nTotal data rows: {ws.max_row - 1}")
