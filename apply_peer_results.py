#!/usr/bin/env python3
"""Apply peer_results.json to Excel and rebuild viewer data."""
import json
from pathlib import Path
from openpyxl import load_workbook

BASE_DIR = Path('/opt/data/home/hermes-arxiv-agent')

# Load results
with open(BASE_DIR / 'peer_results.json') as f:
    results = json.load(f)

wb = load_workbook(BASE_DIR / 'papers_record.xlsx')
ws = wb["Papers"]
headers = [str(c.value or '') for c in ws[1]]
idx = {h: i + 1 for i, h in enumerate(headers)}

# Ensure columns exist
for col_name in ['peer_reviewed', 'peer_venue']:
    if col_name not in idx:
        col = len(headers) + 1
        ws.cell(1, col, col_name)
        headers.append(col_name)
        idx[col_name] = col

count = 0
for row in range(2, ws.max_row + 1):
    sid = str(ws.cell(row, idx['arxiv_id']).value or '').strip()
    if sid in results:
        status, venue = results[sid]
        ws.cell(row, idx['peer_reviewed'], status)
        ws.cell(row, idx['peer_venue'], venue)
        if status:
            count += 1

wb.save(BASE_DIR / 'papers_record.xlsx')
print(f"Applied: {count} peer-reviewed out of {len(results)} checked")
