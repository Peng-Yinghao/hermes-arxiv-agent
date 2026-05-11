#!/usr/bin/env python3
"""Write search results to Excel and update new_papers.json with Chinese summaries."""
import json, openpyxl
from datetime import date
from pathlib import Path
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path(__file__).resolve().parent
EXCEL_FILE = BASE_DIR / 'papers_record.xlsx'

summaries_cn = {
    "2501.11739": "探讨AI智能体中情景记忆的双面性：虽可增强用户对智能体行为的监控与理解，但也引入隐私、操控等新风险。提出四项原则指导情景记忆能力的安全可信发展。",
    "2501.14119": "提出层次化嵌入增强与自主结构化记忆操控框架，通过动态记忆重分配机制优先处理关键上下文特征，在长序列处理和多领域泛化任务中显著提升计算效率与准确性。",
    "2502.16090": "提出Echo模型，创新性地将时间信息融入LLM训练过程，通过多智能体框架生成多轮情景记忆对话数据EM-Train，在EM-Test基准上显著超越现有大模型。",
    "2505.23422": "构建CTIM-Rover软件工程智能体，引入跨任务实例情景记忆。实验表明ExpeL等方法无法扩展至真实SE问题，CTIM项引入的噪声反而导致性能下降。",
    "2503.21760": "提出MemInsight自主记忆增强方法，通过语义数据表示与检索机制优化历史交互，使LLM智能体在对话推荐、问答和摘要任务中准确率提升，推荐说服力提升14%。",
}

with open(BASE_DIR / 'search_results.json', 'r') as f:
    papers = json.load(f)

today = date.today().isoformat()

if EXCEL_FILE.exists():
    wb = openpyxl.load_workbook(EXCEL_FILE)
else:
    wb = openpyxl.Workbook()

if 'Papers' in wb.sheetnames:
    ws = wb['Papers']
    headers = [cell.value for cell in ws[1]]
    existing_ids = {str(row[0]) for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
else:
    ws = wb.active
    ws.title = 'Papers'
    headers = ['arxiv_id', 'title', 'authors', 'published_date', 'categories',
               'summary', 'summary_cn', 'affiliations', 'pdf_url', 'pdf_filename',
               'pdf_local_path', 'crawled_date']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    existing_ids = set()

added = 0
for pid, p in sorted(papers.items()):
    if pid in existing_ids:
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == pid:
                if 'summary_cn' in headers:
                    row[headers.index('summary_cn')].value = summaries_cn.get(pid, '')
                if 'crawled_date' in headers:
                    row[headers.index('crawled_date')].value = today
                break
        print(f"  [{pid}] updated")
        continue

    next_row = ws.max_row + 1
    data = {
        'arxiv_id': pid,
        'title': p['title'],
        'authors': ', '.join(p['authors']),
        'published_date': p['published'],
        'categories': ', '.join(p.get('categories', [])),
        'summary': p['abstract'],
        'summary_cn': summaries_cn.get(pid, ''),
        'affiliations': '',
        'pdf_url': p['pdf_url'],
        'pdf_filename': f"{pid}.pdf",
        'pdf_local_path': p.get('pdf_path', str(BASE_DIR / 'papers' / f'{pid}.pdf')),
        'crawled_date': today,
    }
    for col, h in enumerate(headers, 1):
        ws.cell(row=next_row, column=col, value=data.get(h, ''))
    existing_ids.add(pid)
    added += 1
    print(f"  [{pid}] ADDED")

wb.save(EXCEL_FILE)
print(f"Saved: {added} new, total {len(existing_ids)} papers")

# Update new_papers.json
with open(BASE_DIR / 'new_papers.json', 'r') as f:
    np = json.load(f)
for paper in np['new_papers']:
    paper['summary_cn'] = summaries_cn.get(paper['arxiv_id'], '')
np['pending_count'] = 0
with open(BASE_DIR / 'new_papers.json', 'w', encoding='utf-8') as f:
    json.dump(np, f, ensure_ascii=False, indent=2)
print("new_papers.json updated")
