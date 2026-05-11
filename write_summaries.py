import json
from openpyxl import load_workbook
from datetime import date

with open('/opt/data/home/hermes-arxiv-agent/all_summaries.json', 'r') as f:
    summaries = json.load(f)

wb = load_workbook('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
ws = wb.active

headers = {}
for col_idx, cell in enumerate(ws[1], 1):
    headers[cell.value] = col_idx

arxiv_col = headers.get('arxiv_id')
summary_col = headers.get('summary_cn')
date_col = headers.get('crawled_date')

today = str(date.today())
updated = 0
for row_idx in range(2, ws.max_row + 1):
    aid = str(ws.cell(row=row_idx, column=arxiv_col).value or '')
    if aid in summaries:
        ws.cell(row=row_idx, column=summary_col).value = summaries[aid]
        ws.cell(row=row_idx, column=date_col).value = today
        updated += 1

wb.save('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
print(f'Updated {updated} rows')
