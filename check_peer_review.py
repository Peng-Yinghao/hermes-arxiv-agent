#!/usr/bin/env python3
"""
Detect peer-reviewed papers by checking arXiv journal_ref and comment fields.
Adds peer_reviewed column to Excel and rebuilds viewer data.

Usage:
  python3 check_peer_review.py    # check all papers
"""

from __future__ import annotations

import json
import re
import time
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import load_workbook

import requests

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
ARXIV_API = "https://export.arxiv.org/api/query"

# Known peer-reviewed venues (case-insensitive patterns)
PEER_REVIEW_VENUES = [
    # ML/AI conferences
    "neurips", "nips", "icml", "iclr", "aaai", "ijcai", "aistats", "uai",
    # NLP
    "acl", "emnlp", "naacl", "eacl", "coling", "tacl", "conll",
    # CV/Robotics
    "cvpr", "iccv", "eccv", "icra", "rss", "corl", "iros",
    # Multi-agent
    "aamas", "atal",
    # IR/Web/Data
    "sigir", "www", "kdd", "wsdm", "cikm",
    # Systems/SE
    "icse", "fse", "ase", "osdi", "sosp", "eurosys", "nsdi", "sigcomm",
    # Security
    "ccs", "s&p", "usenix", "ndss",
    # Journals
    "jmlr", "tmlr", "nature", "science", "pnas",
    "ieee transactions", "acm transactions",
    # Generic indicators
    "proceedings of", "accepted at", "published in",
    "conference on", "symposium on", "workshop on",
    "journal of",
]

# Patterns that indicate NOT peer-reviewed (preprints)
NON_PEER_REVIEW_PATTERNS = [
    "preprint", "under review", "submitted to", "arxiv only",
]


def fetch_arxiv_meta(arxiv_id: str) -> dict | None:
    """Fetch paper metadata from arXiv API."""
    url = f"{ARXIV_API}?id_list={arxiv_id}&max_results=1"
    headers = {"User-Agent": "HermesArxivAgent/1.0"}
    try:
        resp = requests.get(url, timeout=15, headers=headers)
        if resp.status_code != 200:
            return None
        ns = {"a": "http://www.w3.org/2005/Atom"}
        root = ET.fromstring(resp.content)
        entry = root.find("a:entry", ns)
        if entry is None:
            return None

        journal_ref = ""
        comment = ""
        jr = entry.find("arxiv:journal_ref", {"arxiv": "http://arxiv.org/schemas/atom"})
        cm = entry.find("arxiv:comment", {"arxiv": "http://arxiv.org/schemas/atom"})
        if jr is not None:
            journal_ref = jr.text or ""
        if cm is not None:
            comment = cm.text or ""

        return {"journal_ref": journal_ref, "comment": comment}
    except Exception:
        return None


def is_peer_reviewed(journal_ref: str, comment: str) -> tuple[bool, str]:
    """Check if paper is peer-reviewed based on arXiv metadata."""
    combined = (journal_ref + " " + comment).lower()

    # Check non-peer-review indicators first
    for pattern in NON_PEER_REVIEW_PATTERNS:
        if pattern in combined:
            return False, ""

    # Check peer-review indicators
    for venue in PEER_REVIEW_VENUES:
        if venue in combined:
            # Extract the venue name from context
            idx = combined.find(venue)
            snippet = combined[max(0, idx - 20):idx + len(venue) + 30]
            return True, snippet.strip()

    return False, ""


def main():
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Papers"]
    headers = [str(c.value or "") for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}

    # Ensure peer_reviewed column exists
    if "peer_reviewed" not in headers:
        col = len(headers) + 1
        ws.cell(row=1, column=col, value="peer_reviewed")
        headers.append("peer_reviewed")
        idx["peer_reviewed"] = col
        print("[INFO] Added peer_reviewed column")

    # Also add peer_venue column
    if "peer_venue" not in headers:
        col = len(headers) + 1
        ws.cell(row=1, column=col, value="peer_venue")
        headers.append("peer_venue")
        idx["peer_venue"] = col

    peer_count = 0
    checked = 0
    total = ws.max_row - 1

    for row in range(2, ws.max_row + 1):
        sid = str(ws.cell(row=row, column=idx["arxiv_id"]).value or "").strip()
        if not sid:
            continue

        # Skip if already checked
        existing = str(ws.cell(row=row, column=idx.get("peer_reviewed", 0)).value or "").strip()
        if existing:
            if "✓" in existing or existing.lower() == "true" or existing.lower() == "yes":
                peer_count += 1
            continue

        checked += 1
        print(f"[{checked}/{total}] Checking {sid}...", end=" ")
        meta = fetch_arxiv_meta(sid)
        time.sleep(1)  # Rate limit

        if meta:
            is_peer, venue = is_peer_reviewed(meta["journal_ref"], meta["comment"])
            status = f"✓ {venue}" if is_peer else ""
            if status:
                ws.cell(row=row, column=idx["peer_reviewed"], value=status)
                ws.cell(row=row, column=idx["peer_venue"], value=venue)
                peer_count += 1
                print(f"PEER: {venue[:60]}")
            else:
                print("preprint")
        else:
            print("API error")

    wb.save(EXCEL_PATH)
    print(f"\n[DONE] {peer_count}/{total} papers are peer-reviewed")
    print(f"Checked {checked} new papers")


if __name__ == "__main__":
    main()
