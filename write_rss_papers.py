#!/usr/bin/env python3
"""Write RSS papers to Excel with Chinese summaries."""
import json, openpyxl
from datetime import date
from pathlib import Path
from openpyxl.styles import Font, PatternFill

BASE_DIR = Path('/opt/data/home/hermes-arxiv-agent')
EXCEL_FILE = BASE_DIR / 'papers_record.xlsx'

summaries_cn = {
    "2605.08060": "发现LLM上下文窗口扩展在多智能体社会困境中反而降低合作意愿，称为记忆诅咒。归因于前瞻性推理意图退化与逐轮激励机制侵蚀，7个LLM、4种游戏、500轮实验验证。",
    "2605.07863": "提出去中心化知识优化框架ADKO，各智能体维护私有高斯过程模型，仅通过紧凑知识令牌通信，实现样本高效、隐私保护的协同黑箱优化。",
    "2605.07692": "提出图加速混合多智能体框架GASim用于大规模社会模拟，通过图结构化记忆检索替代顺序检索，显著降低LLM驱动智能体延迟与计算开销。",
    "2605.07313": "提出规模条件化评估协议，固定任务证据逐步增加无关会话，测量记忆智能体四项诊断指标——预算消耗、检索质量、推理稳定性与污染率。",
    "2605.07242": "形式化智能体记忆的级联更新问题：源数据变更导致派生摘要、缓存陈旧。提出MemoRepair障碍优先级联修复策略，维护记忆派生状态一致性。",
    "2605.07214": "提出异构多智能体协同进化框架HMACE，突破单智能体模板化局限，通过记忆引导探索避免局部最优，实现组合优化启发式自主协同设计。",
    "2605.07180": "提出BoundaryRouter免训练路由框架，冷启动条件下判断查询应使用轻量LLM还是完整智能体执行，利用早期经验学习路由策略降低延迟成本。",
    "2605.07129": "提出排序驱动协同与元记忆双通道检索框架RRCM，融合行为证据与物品元数据，解决LLM推荐中异构证据的上下文构建挑战。",
    "2605.07110": "提出统一架构-生命周期框架保障计算机使用智能体安全，覆盖感知、规划、记忆使用、工具中介、权限与运行时监督六维度，超越任务成功评估范式。",
    "2605.07103": "提出智能体框架ARMOR用于化学反应可行性预测，自适应效用感知多工具推理，动态选择最优工具组合解决单一工具覆盖面不足问题。",
    "2605.07042": "将智能体上下文搜索建模为POMDP，解决大规模环境下工作记忆退化问题，提供结构化搜索基础设施防止冗余探索与重复循环。",
    "2605.06924": "提出智能体自回归扩散架构A²RD用于长视频合成，通过检索-合成-精炼-更新闭环分离创意合成与一致性维护，三智能体协同管理记忆与生成。",
    "2605.06812": "提出统一图表示方法解决LLM智能体安全审计难题，将工具调用、记忆管理、多智能体协作建模为可追溯图结构，弥合底层事件与高层意图的语义鸿沟。",
    "2605.06731": "发现并形式化无意识长期状态中毒风险：个性化智能体日常对话逐渐弱化确认边界、扩大工具默认范围、升级自主行为，提出系统化防御方案。",
    "2605.06716": "综述LLM智能体记忆机制从存储到经验的演变，提出操作系统工程与认知科学统一视角，梳理记忆在智能体架构中的核心作用与未来方向。",
    "2605.06702": "将部署时学习形式化为LLM生命周期第三阶段，提出案例持续适应方法CASCADE，使智能体从交互经验中持续学习，打破训练与部署的刚性分离。",
    "2605.05583": "提出信念记忆机制解决部分可观测下的智能体记忆问题，用概率信念替代确定性存储，避免过早确定性结论导致的自增强错误与鲁棒性下降。",
    "2604.24372": "提出策略空间进化方法SeaEvo用于自动算法发现，将自然语言推理组织为持久群体策略状态而非瞬态变异上下文，引导进化搜索沿有前景方向前进。",
    "2604.23938": "提出多智能体框架TSAssistant支持药物靶点安全评估自动报告，模块化分段生成与人机协同设计，整合遗传、转录组、药理等多源证据。",
    "2511.02805": "提出MemSearcher框架，端到端强化学习训练LLM在多轮搜索中维护紧凑记忆，仅保留问题相关信息稳定上下文长度，降低计算开销。",
    "2508.15294": "提出多记忆分段系统，模拟人类多维多组件记忆生成过程，超越简单摘要式记忆存储，从多维度生成高质量长期记忆内容填补研究空白。",
    "2506.04565": "综述复合AI系统范式，整合LLM与检索器、智能体、工具、编排器等外部组件，克服独立模型在记忆、推理、实时接地与多模态理解方面的局限。",
}

with open(BASE_DIR / 'new_to_add.json') as f:
    papers = json.load(f)

today = date.today().isoformat()

# Load or create Excel
if EXCEL_FILE.exists():
    wb = openpyxl.load_workbook(EXCEL_FILE)
else:
    wb = openpyxl.Workbook()

if 'Papers' in wb.sheetnames:
    ws = wb['Papers']
else:
    ws = wb.active
    ws.title = 'Papers'
    headers = ['arxiv_id', 'title', 'authors', 'published_date', 'categories',
               'summary', 'summary_cn', 'affiliations', 'pdf_url', 'pdf_filename',
               'pdf_local_path', 'crawled_date']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

# Get existing IDs and headers
headers = [cell.value for cell in ws[1]]
existing_ids = {str(row[0].value) for row in ws.iter_rows(min_row=2) if row[0].value}

# Map S2 category to arxiv-style category
def map_category(cats):
    return ', '.join(cats) if cats else ''

added = 0
for pid, p in sorted(papers.items()):
    if pid in existing_ids:
        print(f"  [{pid}] already in Excel, skipping")
        continue

    next_row = ws.max_row + 1
    data = {
        'arxiv_id': pid,
        'title': p['title'],
        'authors': p['authors_raw'],
        'published_date': p.get('pub_date', '').split(' ')[0] if p.get('pub_date') else '',
        'categories': map_category(p.get('categories', [])),
        'summary': p['abstract'],
        'summary_cn': summaries_cn.get(pid, ''),
        'affiliations': '',
        'pdf_url': p['pdf_url'],
        'pdf_filename': f"{pid}.pdf",
        'pdf_local_path': str(BASE_DIR / 'papers' / f'{pid}.pdf'),
        'crawled_date': today,
    }
    for col, h in enumerate(headers, 1):
        ws.cell(row=next_row, column=col, value=data.get(h, ''))
    added += 1
    print(f"  [{pid}] ADDED: {p['title'][:60]}")

wb.save(EXCEL_FILE)
print(f"\nSaved: {added} new papers, total rows: {ws.max_row}")

# Also add to crawled_ids.txt
with open(BASE_DIR / 'crawled_ids.txt', 'a') as f:
    for pid in sorted(papers.keys()):
        if pid not in existing_ids:
            f.write(f"{pid}\n")
print(f"Added to crawled_ids.txt")
