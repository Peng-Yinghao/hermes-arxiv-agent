import json
from openpyxl import load_workbook

with open('/opt/data/home/hermes-arxiv-agent/all_summaries.json', 'r') as f:
    summaries = json.load(f)

wb = load_workbook('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
ws = wb.active

headers = {}
for col_idx, cell in enumerate(ws[1], 1):
    headers[cell.value] = col_idx

aid_col = headers['arxiv_id']
title_col = headers['title']
authors_col = headers['authors']
published_col = headers['published_date']
summary_col = headers['summary_cn']

papers = []
for row in ws.iter_rows(min_row=2, values_only=True):
    aid = str(row[aid_col - 1]) if row[aid_col - 1] else ''
    if aid in summaries:
        papers.append({
            'arxiv_id': aid,
            'title': str(row[title_col - 1] or ''),
            'authors': str(row[authors_col - 1] or ''),
            'published_date': str(row[published_col - 1] or ''),
            'summary_cn': str(row[summary_col - 1] or ''),
        })

print(f"Found {len(papers)} papers with summaries")

# Print for WeChat message
for i, p in enumerate(papers):
    print(f"PAPER|{i}|{p['arxiv_id']}|{p['title']}|{p['authors']}|{p['published_date']}|{p['summary_cn']}")
