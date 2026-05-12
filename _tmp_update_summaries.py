import openpyxl

summaries = {
    "2605.06746": '本文研究强化学习智能体中的因果涌现现象。作者使用PhiID框架测量智能体潜在表征的因果涌现程度，在多种算法、架构和环境中进行实验。结果提出\u201c因果涌现对齐假说\u201d：成功训练的智能体在训练早期即表现出能够预测最终奖励的因果涌现，其表征动态与奖励提升同步。该发现揭示了因果涌现是RL智能体神经表征重组的新维度，为建立因果干预以改进RL训练提供了新视角，也展现了生物和人工智能体在学习中的深层共性。',
    "2605.07358": '本文系统综述了LLM智能体技能（Agent Skills）的研究现状。作者将技能定义为在任务约束下协调工具、记忆和运行时上下文的可复用程序构件，与智能体的高层推理形成互补。论文围绕技能生命周期的四个阶段\u2014\u2014表征、获取、检索和演化\u2014\u2014组织文献，涵盖代表性方法和应用场景。最后讨论了质量控制、互操作性、安全更新和长期能力管理等开放挑战，并提供了开源资源集合。',
    "2605.07594": '本文指出现有具身智能体记忆系统采用静态的\u201c事前单体记忆注入\u201d（AMMI）范式，易与智能体演化状态失配。作者提出MemCompiler，将记忆利用重构为\u201c状态条件记忆编译\u201d：通过学习的记忆编译器读取当前执行状态，动态选择并编译相关记忆为可执行指导，同时通过文本通道和潜在Soft-Mem通道传递信息。在三个基准上提升高达129%，推理延迟降低60%，证明状态感知的记忆编译兼顾了效果与效率。'
}

wb = openpyxl.load_workbook('papers_record.xlsx')
ws = wb.active

header = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
arxiv_id_col = header.get('arxiv_id')
summary_cn_col = header.get('summary_cn')
crawled_date_col = header.get('crawled_date')

updated = 0
for row in range(2, ws.max_row + 1):
    aid = ws.cell(row=row, column=arxiv_id_col).value
    if aid in summaries:
        ws.cell(row=row, column=summary_cn_col).value = summaries[aid]
        ws.cell(row=row, column=crawled_date_col).value = '2026-05-11'
        updated += 1
        print(f"[OK] Updated {aid}")

wb.save('papers_record.xlsx')
print(f"\n[DONE] Updated {updated} papers in Excel")
