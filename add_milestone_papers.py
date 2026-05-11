#!/usr/bin/env python3
"""Add 7 milestone papers (2023-2024) to papers_record.xlsx with Chinese summaries."""

import json
from datetime import date
from openpyxl import load_workbook
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
EXCEL_PATH = BASE_DIR / "papers_record.xlsx"
MILESTONE_PATH = BASE_DIR / "milestone_papers.json"

# Chinese summaries & classifications for each milestone paper
PAPER_DATA = {
    "2304.03442": {
        "summary_cn": "提出生成式智能体(Generative Agents)架构，扩展LLM以自然语言存储智能体的完整经历、将记忆合成为高层反思、并动态检索以规划行为。在模拟小镇中，25个智能体展现出可信的个体和涌现社交行为（如自主组织情人节派对）。消融实验证明观察-规划-反思三个组件对行为可信性均至关重要。这是Agent Memory领域的奠基之作，引入了Memory Stream概念。",
        "scene": "多智能体模拟",
        "form": "文本",
        "mem_type": "反思经验型",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2303.11366": {
        "summary_cn": "提出Reflexion框架，通过语言反馈而非权重更新来强化语言智能体。智能体对任务反馈信号进行口头反思，将反思文本维护在情景记忆缓冲区中，以改进后续决策。Reflexion支持多种反馈类型（标量值或自由语言）和来源（外部或内部模拟）。在HumanEval编程基准上达到91% pass@1准确率，超过GPT-4的80%。这是反思型Agent Memory的里程碑工作。",
        "scene": "代码Agent",
        "form": "文本",
        "mem_type": "反思经验型",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2310.08560": {
        "summary_cn": "借鉴操作系统的分层内存管理思想，提出虚拟上下文管理技术，通过数据在快速和慢速内存之间的移动来提供超越LLM上下文窗口限制的扩展上下文。MemGPT系统智能管理不同内存层级，利用中断管理控制流。在文档分析（处理远超上下文窗口的大型文档）和多轮对话（创建能记忆、反思、动态演进的对话智能体）两个领域验证了有效性。这是分层记忆管理的开创性工作。",
        "scene": "个人助手",
        "form": "混合/其他",
        "mem_type": "分层管理型",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2305.16291": {
        "summary_cn": "提出Voyager，首个LLM驱动的Minecraft具身终身学习智能体，无需人类干预即可持续探索世界、获取多样化技能并做出新发现。三大核心组件：自动课程（最大化探索）、不断增长的技能库（存储和检索可执行代码）、迭代提示机制（整合环境反馈和自验证）。Voyager获得3.3倍独特物品、行进2.3倍距离、解锁科技树里程碑快15.3倍。技能库支持跨世界泛化。",
        "scene": "具身Agent",
        "form": "结构化",
        "mem_type": "检索型",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2309.02427": {
        "summary_cn": "借鉴认知科学和符号AI的丰富历史，提出语言智能体的认知架构(CoALA)。CoALA描述了具有模块化记忆组件、结构化动作空间（与内部记忆和外部环境交互）以及通用决策过程的语言智能体。论文回顾性地梳理了大量近期工作，并前瞻性地识别出通往更强大智能体的可行方向。CoALA将当今语言智能体置于更广阔的AI历史背景中。",
        "scene": "通用/其他",
        "form": "混合/其他",
        "mem_type": "分层管理型",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2404.13501": {
        "summary_cn": "首篇系统综述LLM智能体记忆机制。讨论了LLM智能体中记忆的\"是什么\"和\"为什么需要\"，系统回顾了记忆模块的设计和评估方法。提出了记忆机制的分类体系，涵盖记忆的来源、形式、操作和更新策略。同时介绍了记忆模块发挥重要作用的智能体应用场景。创建了持续更新的GitHub仓库跟踪该领域最新进展。",
        "scene": "通用/其他",
        "form": "文本",
        "mem_type": "其他",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
    "2402.02716": {
        "summary_cn": "首篇系统综述LLM智能体规划能力。提出了LLM智能体规划的分类体系：任务分解、计划选择、外部模块、反思与记忆。对每个方向进行了全面分析，讨论了该领域面临的挑战。该综述将记忆作为LLM智能体规划的关键组成部分加以分析。",
        "scene": "通用/其他",
        "form": "文本",
        "mem_type": "其他",
        "peer_reviewed": "否",
        "peer_venue": "",
    },
}

def main():
    with open(MILESTONE_PATH) as f:
        papers = json.load(f)

    wb = load_workbook(EXCEL_PATH)
    ws = wb["Papers"]

    # Get headers
    headers = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    print(f"Headers: {len(headers)} columns")

    today = date.today().isoformat()

    added = 0
    for paper in papers:
        aid = paper["arxiv_id"]
        extra = PAPER_DATA.get(aid)
        if not extra:
            print(f"[SKIP] {aid}: no summary data")
            continue

        # Build row values
        row = []
        for h in headers:
            if h == "arxiv_id":
                row.append(aid)
            elif h == "title":
                row.append(paper["title"])
            elif h == "authors":
                row.append(paper["authors"])
            elif h == "affiliations":
                row.append("")
            elif h == "published_date":
                row.append(paper["published"])
            elif h == "categories":
                row.append(paper["categories"])
            elif h == "abstract":
                row.append(paper["abstract"])
            elif h == "summary_cn":
                row.append(extra["summary_cn"])
            elif h == "pdf_filename":
                row.append("")
            elif h == "crawled_date":
                row.append(today)
            elif h == "notes":
                row.append("[Milestone] Foundational paper in Agent Memory research")
            elif h in ("", None) or h.startswith("_"):
                row.append("")
            elif h == "pdf_url":
                row.append(paper["pdf_url"])
            elif h == "pdf_local_path":
                row.append("")
            elif h == "scene":
                row.append(extra["scene"])
            elif h == "form":
                row.append(extra["form"])
            elif h == "mem_type":
                row.append(extra["mem_type"])
            elif h == "peer_reviewed":
                row.append(extra["peer_reviewed"])
            elif h == "peer_venue":
                row.append(extra["peer_venue"])
            else:
                row.append("")

        ws.append(row)
        print(f"[ADDED] {aid}: {paper['title'][:80]}")
        added += 1

    wb.save(EXCEL_PATH)
    print(f"\nDone: added {added} papers to {EXCEL_PATH}")

if __name__ == "__main__":
    main()
