#!/usr/bin/env python3
"""Slowly fetch missing metadata for new papers (background)."""
import json
import sys
import time
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path

BASE_DIR = Path('/opt/data/home/hermes-arxiv-agent')
sys.path.insert(0, str(BASE_DIR))
from openpyxl import load_workbook

NS = {'a': 'http://www.w3.org/2005/Atom'}
ENDPOINTS = [
    'https://export.arxiv.org/api/query',
    'http://export.arxiv.org/api/query',
]

# Load Excel and find papers missing metadata
wb = load_workbook(BASE_DIR / 'papers_record.xlsx')
ws = wb['Papers']
headers = [str(c.value) if c.value else '' for c in ws[1]]
id_col = headers.index('arxiv_id')
abstract_col = headers.index('abstract')
authors_col = headers.index('authors')
cats_col = headers.index('categories')

# Find rows needing metadata
to_fetch = []
for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    aid = str(row[id_col]) if row[id_col] else ''
    abstract = str(row[abstract_col]) if row[abstract_col] else ''
    authors = str(row[authors_col]) if row[authors_col] else ''
    cats = str(row[cats_col]) if row[cats_col] else ''
    
    if not aid:
        continue
    if not abstract or not authors or not cats:
        to_fetch.append((row_idx, aid))
        print(f"  Need metadata [{row_idx}] {aid}: abs={'✗' if not abstract else '✓'} auth={'✗' if not authors else '✓'} cat={'✗' if not cats else '✓'}")

print(f"\nPapers needing metadata: {len(to_fetch)}")

if not to_fetch:
    print("All metadata complete!")
    sys.exit(0)

updated = 0
for row_idx, aid in to_fetch:
    print(f"[{updated+1}/{len(to_fetch)}] Fetching {aid}...")
    
    meta = None
    for ep in ENDPOINTS:
        url = f"{ep}?id_list={aid}"
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'HermesAgent/1.0 (https://github.com/Peng-Yinghao/hermes-arxiv-agent)'
            })
            with urllib.request.urlopen(req, timeout=20) as resp:
                data = resp.read()
            root = ET.fromstring(data)
            entry = root.find('a:entry', NS)
            if entry is not None:
                title_e = entry.find('a:title', NS)
                summary_e = entry.find('a:summary', NS)
                author_es = entry.findall('a:author', NS)
                cat_es = entry.findall('a:category', NS)
                
                meta = {
                    'title': title_e.text.strip().replace('\n', ' ') if title_e is not None and title_e.text else '',
                    'abstract': summary_e.text.strip().replace('\n', ' ') if summary_e is not None and summary_e.text else '',
                    'authors': ', '.join(a.find('a:name', NS).text for a in author_es if a.find('a:name', NS) is not None),
                    'categories': ', '.join(c.get('term') for c in cat_es),
                }
            break
        except Exception as e:
            continue
    
    if meta is None:
        print(f"  ✗ FAILED")
        time.sleep(5)
        continue
    
    # Update Excel row
    ws.cell(row=row_idx, column=abstract_col + 1, value=meta['abstract'][:4000])
    ws.cell(row=row_idx, column=authors_col + 1, value=meta['authors'])
    ws.cell(row=row_idx, column=cats_col + 1, value=meta['categories'])
    updated += 1
    print(f"  ✓ [{aid}] {meta['title'][:70]}")
    
    if updated % 3 == 0:
        wb.save(BASE_DIR / 'papers_record.xlsx')
        print(f"  --- saved at {updated} ---")
    
    time.sleep(5)

wb.save(BASE_DIR / 'papers_record.xlsx')
print(f"\n✓ Complete! Updated {updated}/{len(to_fetch)} papers")
