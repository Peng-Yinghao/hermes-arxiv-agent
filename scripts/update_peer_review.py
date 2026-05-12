#!/usr/bin/env python3
"""Update peer-review status in Excel from Semantic Scholar data."""
import json
import sys
sys.path.insert(0, '/opt/data/home/hermes-arxiv-agent')
from openpyxl import load_workbook

# Load updates from S2 batch query
with open('/tmp/peer_review_updates.json') as f:
    updates = {u['arxiv_id']: u for u in json.load(f)}

# Add TMLR explicitly (was "Trans. Mach. Learn. Res." — abbreviation not caught)
updates['2309.02427'] = {
    'arxiv_id': '2309.02427',
    'venue': 'Transactions on Machine Learning Research (TMLR)',
    'is_peer': True
}
# Also TACL if any
updates['2503.21760'] = {
    'arxiv_id': '2503.21760',
    'venue': 'EMNLP 2025',
    'is_peer': True
}

wb = load_workbook('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
ws = wb['Papers']

# Find column indices from headers
headers = [str(c.value) if c.value else '' for c in ws[1]]
print(f"Headers: {headers}")
id_col = headers.index('arxiv_id')
peer_col = headers.index('peer_reviewed') if 'peer_reviewed' in headers else -1
venue_col = headers.index('peer_venue') if 'peer_venue' in headers else -1

updated = 0
for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
    aid = str(row[id_col].value) if row[id_col].value else ''
    if aid in updates:
        info = updates[aid]
        if info.get('is_peer'):
            if peer_col >= 0:
                ws.cell(row=row_idx, column=peer_col + 1, value='Yes')
            if venue_col >= 0 and info.get('venue'):
                ws.cell(row=row_idx, column=venue_col + 1, value=info['venue'])
            updated += 1
            print(f"  ✓ [{aid}] → {info.get('venue', '')[:60]}")

wb.save('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
print(f"\nUpdated {updated} papers with peer-review status")
