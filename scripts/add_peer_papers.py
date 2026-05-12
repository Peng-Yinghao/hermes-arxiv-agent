#!/usr/bin/env python3
"""Batch add new peer-reviewed papers to the system."""
import os
import sys
import json
import time
import urllib.request
import xml.etree.ElementTree as ET
from datetime import date
from pathlib import Path

BASE_DIR = Path('/opt/data/home/hermes-arxiv-agent')
sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

NS = {'a': 'http://www.w3.org/2005/Atom'}
ENDPOINTS = [
    'https://export.arxiv.org/api/query',
    'http://export.arxiv.org/api/query',
]

# Load new peer-reviewed IDs
with open('/tmp/new_peered_ids.txt') as f:
    NEW_IDS = [line.strip() for line in f if line.strip()]

# Take top 35 (skip irrelevant ones)
TOP_IDS = NEW_IDS[:35]

# Load existing IDs to avoid duplicates
wb = load_workbook(BASE_DIR / 'papers_record.xlsx')
ws = wb['Papers']
headers = [str(c.value) if c.value else '' for c in ws[1]]
id_col = headers.index('arxiv_id')
existing_ids = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[id_col]:
        existing_ids.add(str(row[id_col]).strip())

# Filter out existing
to_fetch = [aid for aid in TOP_IDS if aid not in existing_ids]
print(f"New IDs to fetch: {len(to_fetch)}/{len(TOP_IDS)}")
print(f"Already in Excel: {len(TOP_IDS) - len(to_fetch)}")

# Load S2 venue data
with open('/tmp/s2_citing_papers.json') as f:
    s2_data = json.load(f)

def fetch_arxiv_metadata(arxiv_id):
    """Fetch paper metadata from arXiv API."""
    for ep in ENDPOINTS:
        url = f"{ep}?id_list={arxiv_id}"
        try:
            req = urllib.request.Request(url, headers={
                'User-Agent': 'HermesAgent/1.0 (https://github.com/Peng-Yinghao/hermes-arxiv-agent)'
            })
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = resp.read()
            root = ET.fromstring(data)
            entry = root.find('a:entry', NS)
            if entry is None:
                return None
            
            title = entry.find('a:title', NS).text.strip().replace('\n', ' ')
            raw_id = entry.find('a:id', NS).text.strip()
            full_id = raw_id.split('/abs/')[-1] if '/abs/' in raw_id else raw_id
            published = entry.find('a:published', NS).text[:10]
            authors = ', '.join(a.find('a:name', NS).text for a in entry.findall('a:author', NS))
            summary = entry.find('a:summary', NS).text.strip().replace('\n', ' ')
            cats = ', '.join(c.get('term') for c in entry.findall('a:category', NS))
            
            return {
                'arxiv_id': full_id.split('v')[0],
                'full_id': full_id,
                'title': title,
                'authors': authors,
                'published_date': published,
                'categories': cats,
                'abstract': summary,
            }
        except Exception as e:
            continue
    return None

# Column mapping
col_map = {h: i for i, h in enumerate(headers)}
today = date.today().isoformat()

added = 0
for aid in to_fetch:
    print(f"  Fetching {aid}...")
    meta = fetch_arxiv_metadata(aid)
    if meta is None:
        print(f"    FAILED")
        continue
    
    # Get peer-review info from S2
    s2 = s2_data.get(aid, {})
    venue = s2.get('venue', '')
    
    # Determine peer_reviewed
    peer_reviewed = 'Yes' if venue else ''
    
    # New row
    row_data = {
        'arxiv_id': meta['arxiv_id'],
        'title': meta['title'],
        'authors': meta['authors'],
        'affiliations': '',
        'published_date': meta['published_date'],
        'categories': meta['categories'],
        'abstract': meta['abstract'],
        'summary_cn': '',  # Will be filled by LLM
        'pdf_filename': '',  # Will be filled by monitor
        'crawled_date': today,
        'notes': '',
        'peer_reviewed': peer_reviewed,
        'peer_venue': venue,
        'scene': '',  # Will be classified
        'form': '',
        'mem_type': '',
    }
    
    # Append row
    row_idx = ws.max_row + 1
    for field, val in row_data.items():
        if field in col_map:
            ws.cell(row=row_idx, column=col_map[field] + 1, value=val)
    
    added += 1
    print(f"    ✓ {meta['title'][:80]} | venue: {venue[:50]}")
    time.sleep(1.2)  # Respect arXiv rate limit

wb.save(BASE_DIR / 'papers_record.xlsx')
print(f"\nAdded {added} new papers to Excel")
