#!/usr/bin/env python3
"""Add new peer-reviewed papers to Excel with available data."""
import json
import sys
sys.path.insert(0, '/opt/data/home/hermes-arxiv-agent')
from openpyxl import load_workbook

with open('/tmp/new_rows.json') as f:
    new_rows = json.load(f)

wb = load_workbook('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
ws = wb['Papers']

headers = [str(c.value) if c.value else '' for c in ws[1]]
col_map = {h: i for i, h in enumerate(headers)}

# Check existing IDs
existing_ids = set()
id_col = headers.index('arxiv_id')
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[id_col]:
        existing_ids.add(str(row[id_col]).strip())

added = 0
skipped = 0
for row_data in new_rows:
    aid = row_data['arxiv_id']
    if aid in existing_ids:
        skipped += 1
        continue
    
    row_idx = ws.max_row + 1
    for field, val in row_data.items():
        if field in col_map:
            ws.cell(row=row_idx, column=col_map[field] + 1, value=val if val else '')
    added += 1

wb.save('/opt/data/home/hermes-arxiv-agent/papers_record.xlsx')
print(f"Added {added}, skipped {skipped} (already existed)")
print(f"Total rows: {ws.max_row - 1}")
